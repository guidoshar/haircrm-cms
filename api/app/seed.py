"""一次性数据 seed：解析 价格体系(7)新天地.xlsx，构建分类 / 项目 / 价格 / 步骤 / 套餐 / 会员等级 / 门店 / Banner。
重复运行幂等：基于 slug/unique 字段判定，已存在则 skip。
"""
from __future__ import annotations

import os
import re
import sys

from sqlalchemy.orm import Session
from openpyxl import load_workbook

from .core.config import settings
from .core.security import hash_password
from .core.storage import upload_local_file, ensure_bucket
from .deps import SessionLocal, engine
from .models import (
    Base,
    AdminUser,
    Banner,
    Category,
    Service,
    ServiceStep,
    ServicePrice,
    ServicePackage,
    MemberTier,
    Store,
    SiteConfig,
    MediaAsset,
)


# ============= 业务常量 =============
CATEGORY_DEFS = [
    {"name": "头皮·清", "slug": "scalp-cleanse", "tagline": "净化毛囊 · 通透呼吸", "accent_color": "#86A07A", "icon_key": "Sparkle", "sort_order": 1},
    {"name": "头皮·调", "slug": "scalp-balance", "tagline": "舒压调理 · 平衡油脂", "accent_color": "#9E8A6E", "icon_key": "Drop", "sort_order": 2},
    {"name": "头皮·补", "slug": "scalp-nourish", "tagline": "焕活毛囊 · 滋养发根", "accent_color": "#B8945A", "icon_key": "Leaf", "sort_order": 3},
    {"name": "头皮·养", "slug": "scalp-care", "tagline": "深度焕活 · 长效养护", "accent_color": "#C8A36B", "icon_key": "Heart", "sort_order": 4},
    {"name": "发芯赋活", "slug": "hair-shaft", "tagline": "丝质柔顺 · 由内修护", "accent_color": "#A38C5F", "icon_key": "Star", "sort_order": 5},
    {"name": "臻享疗愈", "slug": "deep-therapy", "tagline": "古法手法 · 身心舒压", "accent_color": "#6B7A5A", "icon_key": "Hand", "sort_order": 6},
]


# 项目 → 分类映射（基于 Excel "项目分类" 列拆分；古法通养在 Excel 中归在 臻享疗愈）
SERVICE_CATEGORY = {
    # 头皮·清
    "愈养洗": "scalp-cleanse",
    "弱酸奢养洗": "scalp-cleanse",
    "纳米酸养": "scalp-cleanse",
    "奢养净护": "scalp-cleanse",
    # 头皮·调
    "纳米酸护": "scalp-balance",
    "头皮卸妆净护": "scalp-balance",
    "去屑止痒净护": "scalp-balance",
    "水油平衡净护": "scalp-balance",
    "姜养舒护": "scalp-balance",
    # 头皮·补
    "发量焕活": "scalp-nourish",
    "育发新生": "scalp-nourish",
    "青丝养护": "scalp-nourish",
    "养黑能量": "scalp-nourish",
    # 头皮·养
    "多效小金瓶": "scalp-care",
    # 发芯赋活
    "角蛋白养发": "hair-shaft",
    "玫瑰养发": "hair-shaft",
    # 臻享疗愈
    "草本沐息头疗": "deep-therapy",
    "古法舒压头疗": "deep-therapy",
    "古法通养": "deep-therapy",
}


# 项目 → 一句话卖点 / 大白话原理 / 价值
SERVICE_COPY = {
    "愈养洗": {
        "summary": "日常基础净护，洗去尘埃与浮油，让头皮先做个深呼吸。",
        "value_md": "- 适合发量浓密、爱出油、想要日常保养的发友\n- 时间紧又想被认真对待的午休时段首选",
        "products_md": "- 诗碧曼基础洗护系列\n- 电动毛囊按摩仪\n- 少量护发精油",
        "principle_md": "先用闻香与梳发让头皮经络放松，再用循环水流二度清洗，配合电动按摩仪松开毛囊口的浮油与皮屑。中温吹干顺发，发尾点上一滴精油 —— 不复杂，但每一步都到位。",
    },
    "弱酸奢养洗": {
        "summary": "弱酸调理 + 颈椎奢护，洗一次像做了一次小型理疗。",
        "value_md": "- 头皮敏感、发丝毛躁的发友\n- 长时间伏案、肩颈僵紧者格外适合",
        "products_md": "- 诗碧曼弱酸奢养系列洗护\n- 颈椎奢护发膜\n- 纳米雾辅助渗透",
        "principle_md": "用弱酸配方把头皮 PH 拉回最舒服的区间，再加一道颈椎奢护：发膜涂匀后用纳米雾静敷，顺势疏通耳后淋巴与颈椎，洗发同时把肩颈一起放松。",
    },
    "纳米酸养": {
        "summary": "草本汤化作纳米雾，钻进毛囊里去净。",
        "value_md": "- 头皮闷重、油脂偏多、清洗后很快又油的发友\n- 想体验深度净化的发友",
        "products_md": "- 诗碧曼纳米酸养系列\n- 草本汤循环冲淋\n- 碳酸泉修护\n- 头部按摩膏",
        "principle_md": "草本汤汽化成纳米级雾粒，颗粒小到能钻入毛孔，把堵在毛囊口的油渣冲松。再用按摩膏配合手法做颈部提拉，纳米雾熏蒸 10 分钟做软化，最后碳酸泉一冲就净 —— 不是简单洗一洗。",
    },
    "奢养净护": {
        "summary": "奢养感的日常护理，温柔不过度。",
        "value_md": "- 想要日常多一点仪式感的发友\n- 适合染烫后头皮微敏的修复期",
        "products_md": "- 诗碧曼奢养系列\n- 头皮按摩膏",
        "principle_md": "用奢养系列做二次清洁，强调按摩节奏与温度感，帮你在 50 分钟里完成一次完整的头皮放松。",
    },
    "纳米酸护": {
        "summary": "酸性调理 + 纳米雾化护理，给头皮一次深度安抚。",
        "value_md": "- 头皮泛红、起痘的发友\n- 染烫后需要深度调理的发友",
        "products_md": "- 诗碧曼纳米酸护\n- 纳米雾",
        "principle_md": "把酸性护理液雾化成微粒，覆盖每一寸头皮，平衡 PH 之外还能软化老废角质，再用专业手法按出舒适感。",
    },
    "头皮卸妆净护": {
        "summary": "发缝细分，把头皮上的彩妆和顽固污垢一寸寸卸掉。",
        "value_md": "- 习惯抹定型/造型品的发友\n- 觉得『洗不干净』的发友",
        "products_md": "- 诗碧曼专业头皮卸妆液\n- 大头棉棒细分清洁",
        "principle_md": "把头皮当脸来卸：分发缝、用棉棒蘸取卸妆液逐区清洁，再做熏蒸软化，二次冲洗 + 按摩放松，最后机喷水护、刮痧、肩颈按摩三连击。",
    },
    "去屑止痒净护": {
        "summary": "针对头皮屑与痒，舒缓 + 维稳同步进行。",
        "value_md": "- 季节交替头屑明显的发友\n- 头皮易痒易敏感的发友",
        "products_md": "- 诗碧曼去屑舒缓液\n- 头皮舒缓刮痧",
        "principle_md": "先清后护：洗发后吹干，再机喷专业去屑舒缓液，刮痧两遍把成分送进头皮，肩颈按摩同步放松。",
    },
    "水油平衡净护": {
        "summary": "T 区出油、U 区干燥的『混油皮头皮』专项。",
        "value_md": "- 早洗晚油、发尾又干的发友\n- 想找回头皮舒适状态的发友",
        "products_md": "- 诗碧曼水油平衡系列\n- 八爪鱼放松梳",
        "principle_md": "区分头皮与发尾用不同节奏处理：头皮二次清洁 + 按摩放松，发尾涂发膜 + 纳米雾静敷，水油同步回归平衡。",
    },
    "姜养舒护": {
        "summary": "古法姜养，温通经络驱寒湿。",
        "value_md": "- 怕冷、肩颈寒、头皮易出油的发友\n- 经期前后头皮敏感的女士",
        "products_md": "- 诗碧曼姜养泥\n- 草本熏蒸\n- 头部刮痧板",
        "principle_md": "用棉棒蘸姜泥分区按揉头皮，借姜的辛温通开毛孔与经络，熏蒸 10 分钟透进去，再清洗按摩，最后机喷水护与肩颈舒压。",
    },
    "发量焕活": {
        "summary": "30 分钟轻量护理，集中给毛囊补能量。",
        "value_md": "- 觉得最近发量『细一点、塌一点』的发友\n- 时间紧、想短时间见效的发友",
        "products_md": "- 诗碧曼焕活精华\n- 浴帽热敷",
        "principle_md": "无洗发版本，直接上 10ML 焕活精华按摩 25 分钟，戴浴帽吸收并做肩颈舒压 —— 让成分有时间真正进入毛囊。",
    },
    "育发新生": {
        "summary": "针对脱发期与产后掉发的强化方案。",
        "value_md": "- 毛囊还在但发根弱的发友\n- 产后/压力期掉发增多的发友",
        "products_md": "- 诗碧曼育发液（每次 20ML）\n- 头皮刮痧板",
        "principle_md": "洗后吹干，机喷育发液 20ML，戴浴帽让毛囊充分吸收，刮痧把成分往深处推，肩颈按摩放松交感神经 —— 给毛囊一个能呼吸的环境。",
    },
    "青丝养护": {
        "summary": "30 分钟温和养发，给发丝一份基础滋养。",
        "value_md": "- 染烫损伤需要温和修复的发友\n- 想做『加餐』基础养发的发友",
        "products_md": "- 诗碧曼青丝养护精华",
        "principle_md": "无洗发，直接上养护精华 25 分钟 + 浴帽吸收 + 肩颈舒压，节奏温和。",
    },
    "养黑能量": {
        "summary": "针对白发期的能量补给，从根开始养黑。",
        "value_md": "- 早生白发的发友\n- 染发频次高、想减少染发频率的发友",
        "products_md": "- 诗碧曼养黑能量液\n- 头皮刮痧",
        "principle_md": "机喷养黑能量液 20ML，刮痧两遍 + 浴帽吸收 + 肩颈按摩，配合规律护理周期，让发根从内部回到健康状态。",
    },
    "多效小金瓶": {
        "summary": "店内王牌，多效合一的深度修复。",
        "value_md": "- 想一次满足『清+养+护』的发友\n- 长期未做深度护理的发友",
        "products_md": "- 诗碧曼小金瓶（15ML/次）\n- 生发仪",
        "principle_md": "搓热头皮、刮痧提拉、机喷小金瓶 25 分钟，再上生发仪 15 分钟，最后顺发上精油 —— 一次到位的『高密度补给』。",
    },
    "角蛋白养发": {
        "summary": "把流失的角蛋白补回去，发丝重新挺立有光。",
        "value_md": "- 烫染后发丝无弹力的发友\n- 觉得头发『软塌没光泽』的发友",
        "products_md": "- 诗碧曼角蛋白发膜\n- 热蒸养发模式",
        "principle_md": "细分发条涂角蛋白，再用热蒸养发 20 分钟让分子打开发鳞片渗入芯部，凉发后短促冲洗锁住效果。",
    },
    "玫瑰养发": {
        "summary": "玫瑰润发卷圈，整头柔顺有光泽，香气整天都在。",
        "value_md": "- 想要顺、亮、香气感的发友\n- 拍照、约会、出席场合前最佳准备",
        "products_md": "- 诗碧曼玫瑰润发乳\n- 热蒸养发模式",
        "principle_md": "细发条涂上玫瑰润发乳卷圈静置，热蒸 20 分钟让发丝打开纤维结构吸收滋养，再凉发短冲，顺发用精油收尾 —— 柔顺度肉眼可见。",
    },
    "草本沐息头疗": {
        "summary": "草本香气配合手法，做一次清醒的『小睡』。",
        "value_md": "- 失眠、入睡难的发友\n- 长期高压、想要『放空 50 分钟』的发友",
        "products_md": "- 诗碧曼草本头疗膏\n- 古法手法",
        "principle_md": "以草本香开启嗅觉放松，配合古法头疗专业手法 40 分钟带你进入头皮深度休息，发丝点上一滴精油结束，整个人像重启过一样。",
    },
    "古法舒压头疗": {
        "summary": "店内招牌古法，60 分钟头脑清明。",
        "value_md": "- 高强度脑力工作者\n- 偏头痛、紧张性头痛的发友",
        "products_md": "- 诗碧曼古法专业护理\n- 古法手法",
        "principle_md": "全程古法手法 40 分钟舒压、放松、辨证，结合中医视角观察头皮状态，结束有 5 分钟『古法辨证』小结，让你知道自己头皮在哪个阶段。",
    },
    "古法通养": {
        "summary": "头疗 + 背部疏通双重组合，整身通起来。",
        "value_md": "- 长期肩颈僵硬、背紧不易松的发友\n- 想要『一次彻底放松』的发友",
        "products_md": "- 诗碧曼古法系列\n- 古法手法 + 背部疏通",
        "principle_md": "古法头疗 40 分钟 + 背部古法疏通 30 分钟，从头到背走一遍经络，再结合古法辨证总结当次状态。一次 90 分钟的『慢信式』养护。",
    },
}


# 体验价（从 Excel 第 8 列）。Excel 数据已包含：
TASTE_PRICES = {
    "愈养洗": 98, "弱酸奢养洗": 147, "纳米酸养": 228, "奢养净护": 198,
    "纳米酸护": 298, "头皮卸妆净护": 158, "去屑止痒净护": 168, "水油平衡净护": 188, "姜养舒护": 208,
    "发量焕活": 138, "育发新生": 268, "青丝养护": 168, "养黑能量": 288,
    "多效小金瓶": 398,
    "角蛋白养发": 238, "玫瑰养发": 268,
    "草本沐息头疗": 218, "古法舒压头疗": 358, "古法通养": 468,
}


def slugify(name: str) -> str:
    return (
        name.replace("·", "-")
        .replace(" ", "-")
        .replace("（", "")
        .replace("）", "")
        .lower()
    )


# ============= Excel 解析 =============
def parse_workbook(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    ws_price = wb["价格体系"]
    ws_steps = wb["操作步骤"]

    rows = list(ws_price.iter_rows(values_only=True))
    services_data = {}  # name -> dict
    current_name = None
    for i, row in enumerate(rows):
        if i < 2:
            continue
        col_b = row[1]
        col_c = row[2]
        col_d = row[3]
        col_e = row[4]
        col_f = row[5]
        col_g = row[6]

        if col_b and isinstance(col_b, str) and col_b.strip() and isinstance(col_c, (int, float)):
            current_name = col_b.strip()
            services_data.setdefault(current_name, {})
            services_data[current_name].update(
                {
                    "store_price": int(col_c) if col_c else 0,
                    "member_price": int(col_d) if col_d else 0,
                    "platinum_price": int(col_e) if col_e else 0,
                    "diamond_price": int(col_f) if col_f else 0,
                    "time_min": int(col_g) if col_g else 0,
                    "packages": services_data.get(current_name, {}).get("packages", []),
                }
            )
        elif current_name:
            for cell in (col_c, col_f):
                if cell and isinstance(cell, str):
                    s = cell.strip()
                    if not s:
                        continue
                    pkg = parse_package_text(s)
                    if pkg:
                        services_data[current_name].setdefault("packages", []).append(pkg)

    addon_specials = {}
    for row in rows:
        col_b = row[1]
        col_c = row[2]
        col_g = row[6]
        if col_b in ("肩颈减压按摩", "眼部减压按摩"):
            addon_specials[col_b] = {
                "label": "叠加项",
                "price_text": str(col_c) if col_c else "",
                "time": str(col_g) if col_g else "",
            }

    steps_by_service: dict[str, list[dict]] = {}
    rows_s = list(ws_steps.iter_rows(values_only=True))
    cur_service = None
    cur_minutes = 0
    cur_step_text = []
    for row in rows_s:
        a = row[0]
        b = row[1]
        c = row[2]
        if isinstance(b, str) and b.strip() and b.strip() not in ("项目名称", "步骤"):
            if cur_service and cur_step_text:
                steps_by_service.setdefault(cur_service, []).append(
                    {"minutes": cur_minutes, "text": "\n".join(cur_step_text).strip()}
                )
            cur_service = b.strip()
            cur_minutes = int(c) if isinstance(c, (int, float)) else 0
            cur_step_text = []
            steps_by_service.setdefault(cur_service, [])
            continue
        if isinstance(a, str) and a.strip() == "步骤" and isinstance(b, str) and b.strip():
            cur_step_text.append(b.strip())
        elif b and isinstance(b, str) and b.strip() and not c:
            if cur_service and cur_step_text:
                cur_step_text.append(b.strip())
    if cur_service and cur_step_text:
        steps_by_service.setdefault(cur_service, []).append(
            {"minutes": cur_minutes, "text": "\n".join(cur_step_text).strip()}
        )

    return services_data, steps_by_service


def parse_package_text(s: str) -> dict | None:
    """解析诸如 '课程价：498元/5次' / '990元/10送1' / '季卡18次：4960元/18送2' / '半年卡36次：9900元/36送6'"""
    s = s.strip()
    if "课程价" in s:
        m = re.search(r"(\d+)\s*元.*?(\d+)\s*次", s)
        if m:
            return {"kind": "course5", "label": "课程价", "price": int(m.group(1)), "times": int(m.group(2)), "gift_count": 0, "options_text": None}
    if "送" in s and "元" in s and "课程价" not in s and "季卡" not in s and "半年" not in s:
        m = re.search(r"(\d+)\s*元/(\d+)送(\d+)", s)
        if m:
            return {
                "kind": "ten1",
                "label": f"{m.group(2)}送{m.group(3)}",
                "price": int(m.group(1)),
                "times": int(m.group(2)),
                "gift_count": int(m.group(3)),
                "options_text": None,
            }
    if "季卡" in s:
        m = re.search(r"季卡(\d+)次.*?(\d+)\s*元/(\d+)送(\d+)", s)
        if m:
            return {
                "kind": "quarter",
                "label": f"季卡 {m.group(1)} 次",
                "price": int(m.group(2)),
                "times": int(m.group(1)),
                "gift_count": int(m.group(4)),
                "options_text": None,
            }
    if "半年卡" in s:
        m = re.search(r"半年卡(\d+)次.*?(\d+)\s*元/(\d+)送(\d+)", s)
        if m:
            return {
                "kind": "half",
                "label": f"半年卡 {m.group(1)} 次",
                "price": int(m.group(2)),
                "times": int(m.group(1)),
                "gift_count": int(m.group(4)),
                "options_text": None,
            }
    if re.search(r"选项|可选", s):
        return {
            "kind": "addon",
            "label": "可选搭配",
            "price": 0,
            "times": 0,
            "gift_count": 0,
            "options_text": s,
        }
    return None


def split_steps(text: str) -> list[dict]:
    """把"1、护理前检测3分钟+复检2分钟；2、闻香..."拆成 step 数组。"""
    if not text:
        return []
    parts = re.split(r"[；;]\s*\d+[、.\s]", text)
    out = []
    for i, p in enumerate(parts):
        p = p.strip()
        if not p:
            continue
        p = re.sub(r"^\d+[、.\s]+", "", p).strip()
        m = re.search(r"(\d+)\s*分", p)
        minutes = int(m.group(1)) if m else 0
        title = p.split("，")[0].split(",")[0].split("（")[0][:24]
        out.append({"idx": i, "title": title or f"步骤{i+1}", "minutes": minutes, "description": p})
    return out


# ============= Tier / Store / Banner =============
TIER_DEFS = [
    {
        "name": "贵宾卡",
        "slug": "vip",
        "fee": 1588,
        "discount_text": "享会员价",
        "benefits_md": "- 茶水 + 点心\n- 品牌洗护用品\n- 生日月礼遇\n- 闺蜜同行券 1 份",
        "accent_color": "#9E8A6E",
        "icon_key": "Sparkle",
        "sort_order": 1,
    },
    {
        "name": "白金卡",
        "slug": "platinum",
        "fee": 3988,
        "discount_text": "7.8 折",
        "benefits_md": "- 品牌功效洗护产品\n- 专属护理包\n- 生日月礼遇\n- 闺蜜同行券 3 份",
        "accent_color": "#B8945A",
        "icon_key": "Crown",
        "sort_order": 2,
    },
    {
        "name": "钻石卡",
        "slug": "diamond",
        "fee": 7988,
        "discount_text": "6.8 折",
        "benefits_md": "- 品牌功效洗护产品\n- 新品免费体验\n- 私人定制护理用品\n- 生日月专属沙龙\n- 闺蜜同行券 6 份\n- 原价卡扣产品",
        "accent_color": "#7A8E68",
        "icon_key": "Diamond",
        "sort_order": 3,
    },
]

STORE_DEFS = [
    {
        "name": "陆家嘴尚悦湾旗舰店",
        "slug": "lujiazui",
        "is_flagship": True,
        "address": "上海市浦东新区陆家嘴尚悦湾",
        "phone": "021-0000-0000",
        "hours": "10:00 - 22:00",
        "intro_md": "旗舰店位于陆家嘴 CBD 核心，玻璃幕墙之间留出一方草本静地，提供完整的诗碧曼养发体验。",
        "sort_order": 1,
    },
    {
        "name": "新天地店",
        "slug": "xintiandi",
        "address": "上海市黄浦区新天地",
        "phone": "021-0000-0000",
        "hours": "10:00 - 22:00",
        "intro_md": "新天地百年石库门内的养发会所，散场前一小时的小憩之所。",
        "sort_order": 2,
    },
    {
        "name": "曹路花园城店",
        "slug": "caolu",
        "address": "上海市浦东新区曹路花园城",
        "phone": "021-0000-0000",
        "hours": "10:00 - 22:00",
        "intro_md": "社区型养发会所，工作日午后小坐，周末与家人同行皆宜。",
        "sort_order": 3,
    },
    {
        "name": "金茂大厦黑金店",
        "slug": "jinmao",
        "is_flagship": True,
        "address": "上海市浦东新区世纪大道 88 号金茂大厦",
        "phone": "021-0000-0000",
        "hours": "10:00 - 22:00",
        "intro_md": "金茂大厦黑金主题店，深色木质与黄铜元素，专为高净值发友定制的私享空间。",
        "sort_order": 4,
    },
]


def seed_admin_user(db: Session):
    if db.query(AdminUser).first():
        return
    u = AdminUser(
        username=settings.ADMIN_USERNAME,
        password_hash=hash_password(settings.ADMIN_PASSWORD),
        display_name="系统管理员",
        is_active=True,
    )
    db.add(u)
    db.commit()
    print(f"[seed] admin user created: {settings.ADMIN_USERNAME} / (env: ADMIN_PASSWORD)")


def seed_site_config(db: Session):
    if db.query(SiteConfig).first():
        return
    cfg = SiteConfig(
        brand_name="诗碧曼·养发会所",
        slogan_main="草本精华  缕缕用心",
        slogan_sub="诗碧曼养发，让你持久年轻",
        intro_md="""我们做的事很简单：用草本配方与古法手法，让你的头皮先放松、再被认真照料。
没有夸张承诺，只在每一寸发缝、每一次按压里把"用心"写下来。
""",
        principle_md="""**草本精华 · 缕缕用心**
- 草本配方：姜、玫瑰、艾草等天然植物原料，按人不同状态调配
- 古法手法：辨证施护，结合中医视角看头皮、耳后淋巴与颈椎
- 数字检测：护理前 + 复检，每一次有据可查的变化
""",
        footer_quote="请尽情享受城市中的小憩时刻，让头发恢复活力。",
        cta_phone="",
        kaiti_enabled=True,
    )
    db.add(cfg)
    db.commit()


def seed_categories(db: Session) -> dict[str, Category]:
    out = {}
    for c in CATEGORY_DEFS:
        cat = db.query(Category).filter(Category.slug == c["slug"]).first()
        if not cat:
            cat = Category(**c)
            db.add(cat)
            db.flush()
        out[c["slug"]] = cat
    db.commit()
    return out


def seed_tiers(db: Session):
    for t in TIER_DEFS:
        if not db.query(MemberTier).filter(MemberTier.slug == t["slug"]).first():
            db.add(MemberTier(**t))
    db.commit()


def seed_stores(db: Session):
    for s in STORE_DEFS:
        if not db.query(Store).filter(Store.slug == s["slug"]).first():
            db.add(Store(**s))
    db.commit()


def seed_services(db: Session, cats: dict[str, Category]):
    services_data, steps_by_service = parse_workbook(settings.SEED_XLSX_PATH)

    sort_per_cat: dict[str, int] = {k: 0 for k in cats}

    for name, data in services_data.items():
        cat_slug = SERVICE_CATEGORY.get(name)
        if not cat_slug:
            continue
        cat = cats[cat_slug]
        slug = slugify(name)
        existing = db.query(Service).filter(Service.slug == slug).first()
        if existing:
            continue

        copy = SERVICE_COPY.get(name, {})
        sort_per_cat[cat_slug] += 1
        s = Service(
            category_id=cat.id,
            name=name,
            slug=slug,
            summary=copy.get("summary", ""),
            principle_md=copy.get("principle_md", ""),
            value_md=copy.get("value_md", ""),
            products_md=copy.get("products_md", ""),
            time_min=data.get("time_min", 0),
            sort_order=sort_per_cat[cat_slug] * 10,
            is_active=True,
        )
        db.add(s)
        db.flush()

        db.add(
            ServicePrice(
                service_id=s.id,
                store_price=data.get("store_price", 0),
                member_price=data.get("member_price", 0),
                platinum_price=data.get("platinum_price", 0),
                diamond_price=data.get("diamond_price", 0),
                taste_price=TASTE_PRICES.get(name, 0),
            )
        )

        steps_blob = steps_by_service.get(name, [])
        if steps_blob:
            text = steps_blob[0].get("text", "")
            steps = split_steps(text)
            for st in steps:
                db.add(ServiceStep(service_id=s.id, **st))

        for i, pkg in enumerate(data.get("packages", [])):
            db.add(ServicePackage(service_id=s.id, sort_order=i, **pkg))

    db.commit()


def seed_banners_and_images(db: Session):
    """把 /seed/img_store/ 下 6 张图作为初始 banner / 服务封面 / 门店配图。"""
    img_dir = settings.SEED_IMG_DIR
    if not os.path.isdir(img_dir):
        print(f"[seed] img dir not found: {img_dir}")
        return
    files = sorted([os.path.join(img_dir, f) for f in os.listdir(img_dir) if f.lower().endswith((".webp", ".jpg", ".jpeg", ".png"))])
    if not files:
        return

    media_keys: list[str] = []
    for path in files:
        try:
            info = upload_local_file(path)
            db.add(
                MediaAsset(
                    key=info["key"],
                    thumb_key=info["thumb_key"],
                    mime=info["mime"],
                    width=info["width"],
                    height=info["height"],
                    bytes=info["bytes"],
                    alt=os.path.basename(path),
                    tag="seed",
                )
            )
            media_keys.append(info["key"])
        except Exception as e:
            print(f"[seed] upload {path} failed: {e}")
    db.commit()

    if not media_keys:
        return

    if not db.query(Banner).first():
        db.add_all(
            [
                Banner(
                    position="home",
                    image_key=media_keys[0],
                    headline="草本精华  缕缕用心",
                    subline="诗碧曼养发，让你持久年轻",
                    cta_text="进入项目",
                    cta_link="/services",
                    sort_order=1,
                ),
                Banner(
                    position="home",
                    image_key=media_keys[min(1, len(media_keys) - 1)],
                    headline="请尽情享受城市中的小憩时刻",
                    subline="让头发恢复活力 —— 这是我们能为你做的事",
                    cta_text="了解理念",
                    cta_link="/about",
                    sort_order=2,
                ),
            ]
        )

    services = db.query(Service).order_by(Service.id).all()
    for i, s in enumerate(services):
        if not s.cover_image_key:
            s.cover_image_key = media_keys[i % len(media_keys)]

    stores = db.query(Store).order_by(Store.sort_order).all()
    for i, st in enumerate(stores):
        if not st.image_key:
            st.image_key = media_keys[(i + 2) % len(media_keys)]

    db.commit()


def main():
    ensure_bucket()
    db = SessionLocal()
    try:
        seed_admin_user(db)
        seed_site_config(db)
        cats = seed_categories(db)
        seed_tiers(db)
        seed_stores(db)
        seed_services(db, cats)
        seed_banners_and_images(db)
        print("[seed] done.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
